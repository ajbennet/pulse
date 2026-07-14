"""
Import the "9-Sig TQQQ Tracker" Google Sheet into SQLite.

The dashboard tab is a grid of stacked key/value blocks laid out side by side
(columns A/B, D/E, G/H). This parser walks each block, tracks the current
section header (a row whose value cell is empty), and emits
(section, key, value_text, value_num) rows. The raw CSV is stored alongside so
nothing is ever lost even if the layout changes.

Data is fetched by the agent via the Google Drive tool and handed to
`import_csv` (or read from a saved file); the parser itself is pure.
"""

import csv
import io
import os
import re
from datetime import datetime
from typing import List, Optional, Tuple

from storage.sqlite_store import DEFAULT_DB, SqliteStore

# Column pairs (key_col, value_col) for the three stacked blocks.
_BLOCKS = [(0, 1), (3, 4), (6, 7)]

_NUM_RE = re.compile(r"[$,%\s]")


def _to_num(text: str) -> Optional[float]:
    """Parse '$242,452.30' / '96.15%' / '-$17,591.20' into a float, else None."""
    if text is None:
        return None
    t = text.strip()
    if not t:
        return None
    is_pct = t.endswith("%")
    cleaned = _NUM_RE.sub("", t)
    if cleaned in ("", "-", "+"):
        return None
    try:
        val = float(cleaned)
    except ValueError:
        return None
    return val / 100.0 if is_pct else val


def parse_dashboard(csv_text: str) -> Tuple[List[Tuple[str, str, str, Optional[float]]], str]:
    """
    Return (rows, captured_at) where rows is a list of
    (section, key, value_text, value_num) and captured_at is the sheet's
    reported refresh time (ISO) if present.
    """
    grid = list(csv.reader(io.StringIO(csv_text)))
    rows: List[Tuple[str, str, str, Optional[float]]] = []
    captured_at = ""

    for key_col, val_col in _BLOCKS:
        section = ""
        for r in grid:
            key = r[key_col].strip() if len(r) > key_col else ""
            val = r[val_col].strip() if len(r) > val_col else ""
            if not key:
                continue
            if not val:
                # Row with a label but no value = a section header.
                section = key
                continue
            rows.append((section, key, val, _to_num(val)))
            if key.lower().startswith("last sheet refresh"):
                captured_at = _normalize_ts(val)

    return rows, captured_at


def _normalize_ts(val: str) -> str:
    for fmt in ("%m/%d/%Y %H:%M:%S", "%m/%d/%Y %H:%M", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(val.strip(), fmt).isoformat()
        except ValueError:
            continue
    return val.strip()


def import_csv(csv_text: str, source_id: str = "", source_name: str = "",
               tab: str = "dashboard", db_path: str = DEFAULT_DB) -> dict:
    """Parse and persist a dashboard CSV. Returns a small summary dict."""
    store = SqliteStore(db_path)
    rows, captured_at = parse_dashboard(csv_text)
    imported_at = datetime.now().isoformat(timespec="seconds")
    import_id = store.record_import(
        raw_csv=csv_text, imported_at=imported_at, source_id=source_id,
        source_name=source_name, tab=tab, captured_at=captured_at,
    )
    store.add_metrics(import_id, rows, captured_at=captured_at)
    return {"import_id": import_id, "tab": tab, "metrics": len(rows),
            "captured_at": captured_at, "imported_at": imported_at}


def import_file(path: str, **kwargs) -> dict:
    """Import a saved CSV file (e.g. from data/imports/)."""
    with open(path, "r") as fh:
        return import_csv(fh.read(), **kwargs)


def import_raw(csv_text: str, tab: str, source_id: str = "", source_name: str = "",
               captured_at: str = "", imported_at: str = "", store=None,
               db_path: str = DEFAULT_DB) -> int:
    """
    Import one raw CSV, routing to the right parser by tab (key/value tabs ->
    metrics, everything else -> cells). Used by restore.
    """
    from storage.sqlite_store import SqliteStore
    store = store or SqliteStore(db_path)
    imported_at = imported_at or datetime.now().isoformat(timespec="seconds")
    kv = _is_key_value(tab)
    if kv and not captured_at:
        _, captured_at = parse_dashboard(csv_text)
    import_id = store.record_import(
        raw_csv=csv_text, imported_at=imported_at, source_id=source_id,
        source_name=source_name, tab=tab, captured_at=captured_at,
    )
    if kv:
        rows, _ = parse_dashboard(csv_text)
        store.add_metrics(import_id, rows, captured_at=captured_at)
    else:
        grid = list(csv.reader(io.StringIO(csv_text)))
        store.add_cells(import_id, tab, parse_table(grid))
    return import_id


# Tabs laid out as key/value blocks (parsed into `metrics`); everything else
# is treated as a normal table (parsed into `sheet_cells`). Matched
# case-insensitively (the live CSV import uses the tab name "dashboard").
KEY_VALUE_TABS = {"dashboard", "dashboard-old", "inputs"}


def _is_key_value(tab: str) -> bool:
    return (tab or "").strip().lower() in KEY_VALUE_TABS


def _cell_text(val) -> str:
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.isoformat()
    return str(val)


def _cell_num(val) -> Optional[float]:
    if isinstance(val, bool):
        return 1.0 if val else 0.0
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        return _to_num(val)
    return None


def parse_table(rows: List[tuple]) -> List[Tuple[int, str, str, Optional[float]]]:
    """
    Parse a tabular sheet (header row + data rows) into
    (row_idx, col_name, value_text, value_num) cells. Empty cells are skipped.
    """
    # First row with >=2 non-empty cells is the header.
    header_i = None
    for i, r in enumerate(rows):
        if sum(1 for c in r if c not in (None, "")) >= 2:
            header_i = i
            break
    if header_i is None:
        return []

    header = rows[header_i]
    names = [(_cell_text(h).strip() or f"col{j}") for j, h in enumerate(header)]

    cells = []
    for ri, r in enumerate(rows[header_i + 1:]):
        if all(c in (None, "") for c in r):
            continue
        for j, val in enumerate(r):
            if val in (None, ""):
                continue
            col = names[j] if j < len(names) else f"col{j}"
            cells.append((ri, col, _cell_text(val), _cell_num(val)))
    return cells


def _rows_to_csv(rows: List[tuple]) -> str:
    buf = io.StringIO()
    writer = csv.writer(buf)
    for r in rows:
        writer.writerow([_cell_text(c) for c in r])
    return buf.getvalue()


def import_xlsx(path: str, source_id: str = "", source_name: str = "",
                db_path: str = DEFAULT_DB) -> List[dict]:
    """
    Import EVERY tab of a downloaded .xlsx workbook (requires openpyxl).

    Each sheet is stored raw (lossless) and parsed: key/value tabs into the
    `metrics` table, tabular tabs into the `sheet_cells` table.
    """
    from openpyxl import load_workbook
    from storage.sqlite_store import SqliteStore

    store = SqliteStore(db_path)
    wb = load_workbook(path, data_only=True, read_only=True)
    imported_at = datetime.now().isoformat(timespec="seconds")
    source_name = source_name or os.path.basename(path)
    results = []

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        csv_text = _rows_to_csv(rows)
        kv = _is_key_value(ws.title)
        captured_at = ""
        if kv:
            _, captured_at = parse_dashboard(csv_text)

        import_id = store.record_import(
            raw_csv=csv_text, imported_at=imported_at, source_id=source_id,
            source_name=source_name, tab=ws.title, captured_at=captured_at,
        )

        if kv:
            metric_rows, _ = parse_dashboard(csv_text)
            n = store.add_metrics(import_id, metric_rows, captured_at=captured_at)
            results.append({"tab": ws.title, "kind": "metrics", "count": n})
        else:
            cells = parse_table(rows)
            n = store.add_cells(import_id, ws.title, cells)
            results.append({"tab": ws.title, "kind": "cells", "count": n})

    return results
